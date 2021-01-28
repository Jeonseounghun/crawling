from os import close
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import urllib.request
from openpyxl.styles import PatternFill, Color
import re
from datetime import datetime
import os


datetime.today()

date = datetime.today().strftime("%Y_%m_%d")


if not os.path.exists("./%s" % date):
    os.makedirs("./%s" % date)


#### 엑셀파일 활성화
wb = Workbook()
wbs = wb.active

#### 카테고리 모아오기
f = open("data_list.txt", "r", encoding="utf8")
data_list = f.readlines()
f.close()
i =1
print(data_list)
for url in data_list:
    ###### 홈페이지 화면 상품 추출
    
    url ="%s" % url.strip()
    header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36"}
    res = requests.get(url, headers=header)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "lxml")
    urls = soup.find_all("a", attrs={"class":"image-container offer-image"})
    name = soup.find("div", attrs={"class":"ranklist-title"}).get_text().strip()
    

    #### 엑셀 칼럼
    wbs.cell(row=1, column=1, value="번호")
    wbs.cell(row=1, column=2, value="상품명")
    wbs.cell(row=1, column=3, value="가격")
    wbs.cell(row=1, column=4, value="이미지")
    wbs.cell(row=1, column=5, value="url")
    wbs.cell(row=1, column=6, value="상세설명")
    

    #### 엑셀 칼럼 스타일 설정
    for j in range(1,16):
        wbs.cell(row=1, column=j).fill = PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid')

    cnt = 1
    for url in urls:
    
        try:
            #### 개별 상품 추출
            url ="%s" % "https:" + url["href"]
            print(url)
            header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36"}
            res = requests.get(url, headers=header)
            res.raise_for_status()
            soup = BeautifulSoup(res.text, "lxml")

            product_name = soup.find("h1", attrs={"class":"ma-title"}).get_text().strip()
            
            img_url = soup.find("a", attrs={"class":"inner dot-app-pd"}).find("img")["data-src"]
            product_detail = soup.find("div", attrs={"class":"do-entry do-entry-separate"}).text.split("  ")

            wbs.cell(row=cnt+1, column=1, value=cnt)
            wbs.cell(row=cnt+1, column=2, value= product_name)
            
            if soup.find("div", attrs={"class":"ma-price-wrap"}):
                price_for_count = soup.find("div", attrs={"class":"ma-price-wrap"}).get_text().strip()
                
                wbs.cell(row=cnt+1, column=3, value= price_for_count)
            
            elif soup.find("span", attrs={"class":"ma-ref-price"}):# for i, k in enumerate(price_for_count): # 수량당 가격
                price_for_count = soup.find("span", attrs={"class":"ma-ref-price"})
                wbs.cell(row=cnt+1, column=3, value= price_for_count.get_text().strip())   
            
            wbs.cell(row=cnt+1, column=4, value= "https:"+img_url)
            wbs.cell(row=cnt+1, column=5, value= url)
            
            for i, k in enumerate(product_detail): # 상품 세부정보
                wbs.cell(row=cnt+1, column=6, value= k)
            print("%s개 진행" % cnt)
            cnt += 1
            
            wb.save("./%s/%s.xlsx" % (date,name))
        except:
            pass
    i += 1
       