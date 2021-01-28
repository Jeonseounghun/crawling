from os import close
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import urllib.request
from openpyxl.styles import PatternFill, Color
cnt = 1
f = open("쿠팡.txt", "a", encoding="utf8")

cnt = 1
name = '보드게임1'



#### 엑셀파일 활성화
wb = Workbook()
wbs = wb.active

#### 엑셀 칼럼
wbs.cell(row=1, column=1, value="번호")
wbs.cell(row=1, column=2, value="상품명")
wbs.cell(row=1, column=3, value="가격")
wbs.cell(row=1, column=4, value="브랜드")
wbs.cell(row=1, column=5, value="이미지")
wbs.cell(row=1, column=6, value="URL")

#### 엑셀 칼럼 스타일 설정
for j in range(1,7):
    wbs.cell(row=1, column=j).fill = PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid')



for i in range(1,26):
    base_url = "https://www.coupang.com/"
    url ="https://partners.coupang.com/#affiliate/ws/best/rocket-fresh"
    header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36"}
    res = requests.get(url, headers=header)

    #### 유효한 url인지 검사 >> 잘못되면 작동하지 않음
    res.raise_for_status()
    soup = BeautifulSoup(res.text,'lxml')
    urls = soup.find_all("li", attrs={"class":"search-product"})

    #### 상품정보 추출
    for url in urls:
        url = url.find("a", attrs={"class":"search-product-link"})
        url = base_url + url["href"]

        url ="%s" % url
        header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36"}
        res = requests.get(url, headers=header)
        res.raise_for_status()
        soup = BeautifulSoup(res.text,'lxml')

        product_name = soup.find("h2", attrs={"class":"prod-buy-header__title"}).get_text()
        review_count = soup.find("span", attrs={"class":"count"}).get_text()
        total_price = soup.find("span", attrs={"class":"total-price"}).get_text().strip()
        img = soup.find("div", attrs={"class":"prod-image__item"}).find("img")["data-src"]
        info = soup.find_all("li", attrs={"class":"prod-attr-item"})

        a= "%s." % cnt + product_name
        b= review_count
        c= total_price
          
        e="https:" + img
        
        q=(url+ "\n")
        
        g=("-"*100+ "\n")
        
        wbs.cell(row=cnt+1, column=1, value=cnt)
        wbs.cell(row=cnt+1, column=2, value=a)
        wbs.cell(row=cnt+1, column=3, value=c)
        wbs.cell(row=cnt+1, column=4, value=b)
        wbs.cell(row=cnt+1, column=5, value=e)
        wbs.cell(row=cnt+1, column=6, value=q)
        count = 7
        for info in info:
            d=(info.get_text())
            wbs.cell(row=cnt+1, column=count, value=d)
            count += 1
        wb.save("%s.xlsx" % "쿠팡")
        print(cnt)
        cnt += 1
        

f.close()