from selenium import webdriver
from bs4 import BeautifulSoup
import time
from openpyxl.styles import PatternFill, Color
from openpyxl import Workbook

cnt = 1
lines = []
name = '장난감2'



#### 엑셀파일 활성화
wb = Workbook()
wbs = wb.active

#### 엑셀 칼럼
wbs.cell(row=1, column=1, value="번호")
wbs.cell(row=1, column=2, value="상품명")
wbs.cell(row=1, column=3, value="가격")
wbs.cell(row=1, column=4, value="리뷰")
wbs.cell(row=1, column=5, value="이미지")
wbs.cell(row=1, column=6, value="URL")
wbs.cell(row=1, column=7, value="상세리뷰")

#### 엑셀 칼럼 스타일 설정
for j in range(1,8):
    wbs.cell(row=1, column=j).fill = PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid')


file = open("url.txt", "r", encoding="utf8")
line = file.readlines()
file.close()

#### 상품들의 상세페이지 url 추출
for line in line:
    
    #### Webdriver 실행
    dr = webdriver.Chrome('./chromedriver.exe') 
    
    #### Webdriver에서 접속하는 URL
    dr.get('%s' % line)
    html = dr.page_source
    soup2 = BeautifulSoup(html, 'lxml')

    #### 각 상품의 url 추출
    urls = soup2.find_all("a", attrs={"class":"a-link-normal a-text-normal"})

    for url in urls:
        lines.append("https://www.amazon.com" + url["href"])
    print(lines)
    dr.quit()

#### 반복문으로 url 읽어오기
for url in lines:
    info_dict ={}
    count = 7
    i = 0
    #### 셀레니움으로 구글드라이버를 통해 url 오픈
    dr2 = webdriver.Chrome('./chromedriver.exe') 
    
    dr2.get("%s" % url )
    html = dr2.page_source
    soup = BeautifulSoup(html, 'lxml')
    time.sleep(3)

    #### 결과값 입력하기 && 정보추출하기
    wbs.cell(row=cnt+1, column=1, value=cnt)

    try:
        title = soup.find('span', attrs={'id':'productTitle'})
        wbs.cell(row=cnt+1, column=2, value=title.get_text().strip())
    except:
        print("title error")
        wbs.cell(row=cnt+1, column=2, value="none")
        pass
    
    try:
        price = soup.find('span', attrs={'id':'price_inside_buybox'})
        wbs.cell(row=cnt+1, column=3, value=price.get_text().strip())
    except:
        print("price error")
        wbs.cell(row=cnt+1, column=3, value="none")
        pass

    try:
        review = soup.find("span", attrs={"data-hook":"rating-out-of-text"}).get_text()
        wbs.cell(row=cnt+1, column=4, value=review)
    except:
        print("review error")
        wbs.cell(row=cnt+1, column=4, value="none")
        pass
    
    


    try:
        img = soup.find("div", attrs={"id":"imgTagWrapperId"}).find("img")["src"]
        wbs.cell(row=cnt+1, column=5, value=img)
    except:
        print("img error")
        wbs.cell(row=cnt+1, column=5, value="none")
        pass

    wbs.cell(row=cnt+1, column=6, value=url)

    try:
        dr2.find_element_by_link_text("See all reviews").click()
        time.sleep(3)
        html = dr2.page_source
        soup = BeautifulSoup(html, 'lxml')
        review_text = soup.find_all("span", attrs={"class":"a-size-base review-text review-text-content"})
        for review_text in review_text:
            wbs.cell(row=cnt+1, column=8+i, value=review_text.get_text().strip())
            i += 1
    except:
        print("review_text")
        wbs.cell(row=cnt+1, column=8+i, value="none")
        pass
            

    wb.save("%s.xlsx" % name)
    print("%s/%s" % (cnt,len(lines)))
    cnt += 1
    dr2.quit()
    