from os import close
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import urllib.request
from openpyxl.styles import PatternFill, Color
import re

#### 변수모음
data_list = []
list_par = []
name_1 = "와인 리스트1"

#### 엑셀파일 활성화
wb = Workbook()
wbs = wb.active

#### 엑셀 칼럼
wbs.cell(row=1, column=1, value="번호")
wbs.cell(row=1, column=2, value="상품명(영어)")
wbs.cell(row=1, column=3, value="상품명(한국)")
wbs.cell(row=1, column=4, value="알코올")
wbs.cell(row=1, column=5, value="용량")
wbs.cell(row=1, column=6, value="와인 종류")
wbs.cell(row=1, column=7, value="탄산 분류")
wbs.cell(row=1, column=8, value="당도")
wbs.cell(row=1, column=9, value="포도품종")
wbs.cell(row=1, column=10, value="제조사")
wbs.cell(row=1, column=11, value="원산지")
wbs.cell(row=1, column=12, value="Color")
wbs.cell(row=1, column=13, value="Aroma")
wbs.cell(row=1, column=14, value="Food Matching")
wbs.cell(row=1, column=15, value="Flavor")
wbs.cell(row=1, column=16, value="Awards / Selling Point")

#### 엑셀 칼럼 스타일 설정
for j in range(1,17):
    wbs.cell(row=1, column=j).fill = PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid')

cnt = 1
for i in range(1,741):
    
    url = "http://www.wine.co.kr/wine/%s" % i
    header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36"}
    res = requests.get(url, headers=header)
    
    #### 유효한 url인지 검사 >> 잘못되면 작동하지 않음
    res.raise_for_status()
    
    soup =BeautifulSoup(res.text, "lxml")

    name = soup.find("h1", attrs={"class":"tit"}).text
    info = soup.find("div", attrs={"class":"more_info clfix"}).text.split('    ')
    info_2 = soup.find("div", attrs={"class":"product_info clfix"}).text.split('    ')
    
    
    english_name = re.sub('[^a-zA-Z0-9]',' ',name).strip()
    korea_name = re.sub('[^ ㄱ-ㅣ가-힣]+',' ',name).strip()
    
    for i in range(8):
        info_2[i].strip().split('   ').append("")
    
    wbs.cell(row=cnt+1, column=1, value=cnt)
    
    try: 
        wbs.cell(row=cnt+1, column=2 ,value=english_name)
    except:
        pass
    
    try: 
        wbs.cell(row=cnt+1, column=3, value=korea_name)
    except:
        pass
    
    try: 
        wbs.cell(row=cnt+1, column=4, value=info[1])
    except:
        pass
    
    try:
        wbs.cell(row=cnt+1, column=5, value=info[2])
    except:
        pass
    
    try: 
        wbs.cell(row=cnt+1, column=6, value=info[3])
    except:
        pass
    
    try:
        wbs.cell(row=cnt+1, column=7, value=info[4])
    except:
        pass
    
    try: 
        wbs.cell(row=cnt+1, column=8, value=info[5])
    except:
        pass
    
    try: 
        wbs.cell(row=cnt+1, column=9, value=info_2[0].strip().split('   ')[1])
    except:
        pass
    
    try:
        wbs.cell(row=cnt+1, column=10, value=info_2[1].strip().split('   ')[1])
    except:
        pass
    
    try: 
        wbs.cell(row=cnt+1, column=11, value=info_2[2].strip().split('   ')[1])
    except:
        pass

    try: 
        wbs.cell(row=cnt+1, column=12, value=info_2[3].strip().split('   ')[1])
    except:
        pass
    
    try: 
        wbs.cell(row=cnt+1, column=13, value=info_2[4].strip().split('   ')[1])
    except:
        pass
    
    try: 
        wbs.cell(row=cnt+1, column=14, value=info_2[5].strip().split('   ')[1])
    except:
        pass
    
    try:
        wbs.cell(row=cnt+1, column=15, value=info_2[6].strip().split('   ')[1])
    except:
        pass
    
    try:
        wbs.cell(row=cnt+1, column=16, value=info_2[7].strip().split('   ')[1])
    except:
        pass
    
    print(cnt)
    cnt += 1

    #### name이라는 제목으로 엑셀 저장
    wb.save("%s.xlsx" % name_1)

   
     


