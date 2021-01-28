from os import close
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
import urllib.request
from openpyxl.styles import PatternFill, Color

#### 변수명
size = ""
tag = ""
list = []
product_list = []
name = "플리스뽀글이"
tag = []
i = 1
cnt = 1

#### 엑셀파일 활성화
wb = Workbook()
wbs = wb.active

#### 엑셀 칼럼
wbs.cell(row=1, column=1, value="번호")
wbs.cell(row=1, column=2, value="상품명")
wbs.cell(row=1, column=3, value="가격")
wbs.cell(row=1, column=4, value="브랜드")
wbs.cell(row=1, column=5, value="사이즈")
wbs.cell(row=1, column=6, value="리뷰수")
wbs.cell(row=1, column=7, value="해시태그")
wbs.cell(row=1, column=8, value="이미지링크")
wbs.cell(row=1, column=9, value="상품링크")
wb.save("%s.xlsx" % name)

#### 상품 url이 담겨있는 파일 읽기
f = open("./url_list.txt", "r", encoding="utf8")
url_list = f.readlines()
f.close

f = open("data_list.txt", "w", encoding="utf8")
f.write(" ")
f.close()

for url in url_list:
 
    ####### 홈페이지 화면 상품 추출
    url ="%s" % url
    header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36"}
    res = requests.get(url, headers=header)
    res.raise_for_status()
    soup = BeautifulSoup(res.text, "lxml")
    
    
    f = open("data_list.txt", "a", encoding="utf8")

    link = soup.find_all("a", {"name":"goods_link"})
    for i in link:
        data = i["href"] + "\n"
        f.write(data)
    f.close()


#### 상품 url이 담겨있는 파일 읽기
f = open("./data_list.txt", "r", encoding="utf8")
data_lists = f.readlines()
f.close


#### 엑셀 칼럼 스타일 설정
for j in range(1,11):
    wbs.cell(row=1, column=j).fill = PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid')


#### 웹 스크랩 작동코드
for data_list in data_lists:
    size = ""
    
    list_2 = []
    #### 데이터 리스트에서 url을 받아 User Agent 설정
    data_list = data_list.strip()
    url ="%s" % data_list
    header = {"User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36"}
    res = requests.get(url, headers=header)
    
    #### 유효한 url인지 검사 >> 잘못되면 작동하지 않음
    res.raise_for_status()

    #### bs4를 이용한 url 설정
    req2 = urllib.request.Request(url)
    res2 = urllib.request.urlopen(url).read()
    soup = BeautifulSoup(res.text, "lxml")
    soup2 = BeautifulSoup(res2,'html.parser')
    soup2 = soup2.find("div",{"class":"product-img"})
    
    #### 중복된 상품 검사
    product = soup.find("span" , attrs={"class":"product_title"})
    if product.get_text().strip() in product_list:
        pass
    
    elif cnt > 0:
        product_list.append(product.get_text().strip())
        
        #### 각 객체값을 호출하는 부분
        price = soup.find("span" , attrs={"class":"product_article_price"})
        review_count = soup.find("a" , attrs={"class":"link_type"})
        brand = soup.find("p", attrs={"class":"product_article_contents"}).find("a")
        sell_cnt = soup.find_all("a", attrs={"class":"listItem"})
        product_size = soup.find_all("div", attrs={"id":"goods_opt_area"})
        imgUrl = soup2.find("img")["src"]
        tag = soup.find_all("a", attrs={"class":"listItem"})
        
        for a in product_size:
            list = a.get_text().split()
        for a in list:
            size += a+","

        for b in range(0,len(tag)):
            list_2.append(tag[b].get_text())

        #### 각 객채값을 엑셀에 입력
        wbs.cell(row=cnt+1, column=1, value=cnt)
        if product:
            wbs.cell(row=cnt+1, column=2, value=product.get_text().strip())
        
        if price:
            wbs.cell(row=cnt+1, column=3, value=price.get_text().strip())
        
        if brand:
            wbs.cell(row=cnt+1, column=4, value=brand.get_text().strip())
    
        # if size:
        #     wbs.cell(row=cnt+1, column=5, value=size)
        
        if review_count:
            wbs.cell(row=cnt+1, column=6, value=review_count.get_text().strip())
        
        wbs.cell(row=cnt+1, column=7, value="")
        wbs.cell(row=cnt+1, column=8, value="https:" + imgUrl)
        wbs.cell(row=cnt+1, column=9, value=url)
        wbs.cell(row=cnt+1, column=10, value=" ")
        
        for i, j in enumerate(list_2):
            wbs.cell(row=cnt+1, column=12+i, value=j)
        
        print("{}개 진행...".format(cnt))
        cnt += 1
        wb.save("%s.xlsx" % name)

    else:
        print("{}개 진행...".format(cnt))
        cnt += 1

        
     

    #### name이라는 제목으로 엑셀 저장