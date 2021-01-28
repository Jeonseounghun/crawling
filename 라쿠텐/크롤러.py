from selenium import webdriver
from bs4 import BeautifulSoup
import time
from openpyxl.styles import PatternFill, Color
from openpyxl import Workbook
import datetime


import requests


date = datetime.datetime.today().strftime("%Y_%m_%d")

cnt = 1
name = '연습'

#### 엑셀파일 활성화
wb = Workbook()
wbs = wb.active

#### 엑셀 칼럼
wbs.cell(row=1, column=1, value="번호")
wbs.cell(row=1, column=2, value="상품명")
wbs.cell(row=1, column=3, value="가격")
wbs.cell(row=1, column=4, value="브랜드")
wbs.cell(row=1, column=5, value="이미지")
wbs.cell(row=1, column=6, value="URL")

#### 엑셀 칼럼 스타일 설정
for j in range(1,7):
    wbs.cell(row=1, column=j).fill = PatternFill(start_color='FFCC00', end_color='FFCC00', fill_type='solid')

#### URL 모음 읽어오기    
f = open("url모음.txt", "r", encoding="utf8")
lines = f.readlines()
f.close
f = open("라쿠텐" + date + ".txt", "a", encoding="utf8")
#### 반복문으로 url 읽어오기
for url in lines:
    
    #### 셀레니움으로 구글드라이버를 통해 url 오픈
    dr2 = webdriver.Chrome('./chromedriver.exe') 
    dr2.get('%s' % url)
    html = dr2.page_source
    soup = BeautifulSoup(html, 'lxml')
    start = datetime.datetime.now()
    end = start + datetime.timedelta(seconds=3)
    while True:
        dr2.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        time.sleep(1)
        if datetime.datetime.now() > end:
            break
    urls1 = soup.find_all("div", attrs={"class":"rnkRanking_top3box rnkRanking_topBgColor"})
    urls2 = soup.find_all("div", attrs={"class":"rnkRanking_top3box"})
    urls3 = soup.find_all("div", attrs={"class":"rnkRanking_after4box"})
    
    urls = urls1 + urls2 + urls3
    dr2.quit()
    
    for url in urls:

        
        product_name = url.find("div", attrs={"class":"rnkRanking_detail"}).get_text().split("\n")
        price = url.find("div", attrs={"class":"rnk_fixedRightBox"}).get_text().strip()
        img = url.find("div", attrs={"class":"rnkRanking_image"}).find("img")["src"]
        url_info = url.find("div", attrs={"class":"rnkRanking_image"}).find("a")["href"]
        a = "%s" % cnt + "\n"
        f.write(a)
        for info in product_name:
            if info:
                b = info+ "\n"
                f.write(b)
        c =  "가격 : " + price+ "\n"       
        f.write(c)
        d = "상품이미지 : " + img+ "\n"
        f.write(d)
        e = "상품상세페이지 : " + url_info+ "\n"
        f.write(e)
        g = "-" *100+ "\n"
        f.write(g)
        
        cnt += 1
    
f.close()


